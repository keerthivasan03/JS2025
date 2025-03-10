import csv
import os
filepath="C:\\Users\\keerthivasan.a\\python programming\\Book_1.csv"
os.path.exists(filepath)
#Dictreader which will read without header
with open(filepath,"r") as read_function:
    read=csv.DictReader(read_function)
    for row in read:
        print("{}, {}, {}".format(row["name"],row["username"],row["department"]))
#read all files
with open(filepath,"r") as read_dict:
    reads=csv.reader(read_dict)
    for row in reads:
        name, username, department=row
        print("{} {} {}".format(name, username, department))
#write
filepath_1="C:\\Users\\keerthivasan.a\\python programming\\Book_9.csv"
os.path.exists(filepath_1)
data = [
    ["Name", " Age", " City"], 
    ["Alice","34", " New York"],
    ["Bob", " 25", " Los Angeles"],
    ["Charlie", " 35", " San Francisco"]]
with open(filepath_1,"a",newline="") as write: # w will write and a will append
    wrote=csv.writer(write)
    wrote.writerows(data)
#write and read
filepath_1="C:\\Users\\keerthivasan.a\\python programming\\Book_9.csv"
os.path.exists(filepath_1)
data = [
    ["Name", " Age", " City"], 
    ["Alice","34", " New York"],
    ["Bob", " 25", " Los Angeles"],
    ["Charlie", " 35", " San Francisco"]]
with open(filepath_1,"a+",newline="") as write: # w will write and a will append
    wrote=csv.writer(write)
    wrote.writerows(data)
    write.seek(0)
    print(write.read())
#write with keys
filepath_2="C:\\Users\\keerthivasan.a\\python programming\\Book_10.csv"
data=[{"Name":"Aisha Khan", "Department":"Engineering", "Salary":"80000"},{"Name":"Jules Lee", "Department":"Marketing", "Salary":"67000"},
      {"Name":"Queenie Corbit", "Department":"Human Resources", "Salary":"90000"}]
keys=["Name","Department","Salary"]
csv.register_dialect("Dialact",delimiter=",",quotechar="~",strict=True, skipinitialspace=True)
with open(filepath_2,"w+", newline="") as writes:
    wroote=csv.DictWriter(writes, dialect="Dialact", fieldnames=keys)
    wroote.writeheader()
    wroote.writerows(data)
    writes.seek(0)
    print(writes.read())
#output print
filelocation="C:\\Users\\keerthivasan.a\\python programming\\contact_info.csv"
with open(filelocation,"r") as update:
    updates=csv.DictReader(update)
    for row in updates:
        print("{} poda dai {} waste {}".format(row["name"],row["role"],row["phone"]))
#append the existing data without header
fileq="C:\\Users\\keerthivasan.a\\python programming\\append2.csv"
csv.register_dialect("Dialacts",delimiter=",",strict=True,skipinitialspace=True)
datas = [{"Name": "Keerthivasan", "Email": "kkk@gmail.com"}]
keys1 = ["Name", "Email"]
with open(fileq,"a",newline="") as no_header:
    header=csv.DictWriter(no_header,fieldnames=keys1,dialect="Dialacts")
    no_header.seek(0)
    if os.stat(fileq).st_size==0:
        header.writeheader()
    header.writerows(datas)
print("data successfully uploaded")
#excel
from openpyxl import load_workbook, Workbook
filepath="C:\\Users\\keerthivasan.a\\python programming\\own_Excel1.xlsx"
def writing():
    wob=Workbook()
    sheet=wob.active #checking whether sheet is active
    sheet.title="Check124"
    data=[["name","Email"],["Kiamia","Kiamia@66.com"],["miakia","miakia@77.com"]]
    for row in data:
        sheet.append(row)
    wob.save(filepath)
    print("Writing done")
def reading():
    reads=load_workbook(filepath)
    sheet_active=reads.active #.active refers to the currently active sheet in the workbook.If the workbook has multiple sheets,active selects the first one by default.
    for row in sheet_active.iter_rows(values_only=True):
        print(row)
    print("reading done")
def append():
    appends=load_workbook(filepath)
    sheet_append=appends.active
    data=[["Rey mysterio","Reymysterio+1@g.com"],["Brocklensar","Brock@g.com"]]
    for rows in data:
        sheet_append.append(rows)
    appends.save(filepath)
    print("Append done")
def delete():
    deletes=load_workbook(filepath)
    row_delete=deletes.active
    row_delete.delete_rows(2)
    deletes.save(filepath)
    print("delete done")
writing()
append()
reading()
delete()
#write with many rows
import csv
import os
print(os.path.exists("C:\\Users\\keerthivasan.a\\python programming\\syslog.log"))
filepath="C:\\Users\\keerthivasan.a\\python programming\\practise_write.csv"
csv.register_dialect("Write",delimiter=",",skipinitialspace=True)
with open(filepath,"w",newline="") as writer_file:
    header = ['Full Name', 'Email Address']
    data = [{'Full Name': 'Blossom Gill', 'Email Address': 'blossom@abc.edu'},
{'Full Name': 'Hayes Delgado', 'Email Address': 'nonummy@utnisia.com'},
{'Full Name': 'Petra Jones', 'Email Address': 'ac@abc.edu'},
{'Full Name': 'Oleg Noel', 'Email Address': 'noel@liberomauris.ca'},
{'Full Name': 'Ahmed Miller', 'Email Address': 'ahmed.miller@nequenonquam.co.uk'},
{'Full Name': 'Macaulay Douglas', 'Email Address': 'mdouglas@abc.edu'},
{'Full Name': 'Aurora Grant', 'Email Address': 'enim.non@abc.edu'},
{'Full Name': 'Madison Mcintosh', 'Email Address': 'mcintosh@nisiaenean.net'},
{'Full Name': 'Montana Powell', 'Email Address': 'montanap@semmagna.org'},
{'Full Name': 'Rogan Robinson', 'Email Address': 'rr.robinson@abc.edu'},
{'Full Name': 'Simon Rivera', 'Email Address': 'sri@abc.edu'},
{'Full Name': 'Benedict Pacheco', 'Email Address': 'bpacheco@abc.edu'},
{'Full Name': 'Maisie Hendrix', 'Email Address': 'mai.hendrix@abc.edu'},
{'Full Name': 'Xaviera Gould', 'Email Address': 'xlg@utnisia.net'},
{'Full Name': 'Oren Rollins', 'Email Address': 'oren@semmagna.com'},
{'Full Name': 'Flavia Santiago', 'Email Address': 'flavia@utnisia.net'},
{'Full Name': 'Jackson Owens', 'Email Address': 'jackowens@abc.edu'},
{'Full Name': 'Britanni Humphrey', 'Email Address': 'britanni@ut.net'},
{'Full Name': 'Kirk Nixon', 'Email Address': 'kirknixon@abc.edu'},
{'Full Name': 'Bree Campbell', 'Email Address': 'breee@utnisia.net'}]
    writer_files=csv.DictWriter(writer_file, dialect="Write",fieldnames=header)
    writer_files.writeheader()
    writer_files.writerows(data)
#Random function testing
import random
def participants(name):
    participants_dict={}
    for names in name:
        participants_dict[names]=random.randint(1,10)
        try:
            if "Keerthi" in participants_dict:
                if participants_dict["Keerthi"]==5:# checking whether it is coming in random
                    return True
                else:
                    raise KeyError("Name not present")
        except KeyError as e:
            print(e)
            return None
name=["Keerthi","anu","isai"]
print(participants(name))
#operator
import operator
fruit = {"oranges": 3, "apples": 5, "bananas": 7, "pears": 2}
#print(sorted(fruit.items(),key=operator.itemgetter(0))) #print(sorted(fruit.items()))
# print(sorted(fruit.items(),key=operator.itemgetter(1)))
#print(sorted(fruit.items(),key=operator.itemgetter(0),reverse=True))
print(sorted(fruit.items(),key=operator.itemgetter(1),reverse=True))

