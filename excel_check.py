from openpyxl import load_workbook, Workbook
filepath="c:\\users\\keerthivasan.a\\own_Excel.xlsx"
def write():
    wob=Workbook()
    sheet=wob.active #create_sheet(title="Bia")
    sheet.title="Own"
    data=[["name","Email"],["Kiamia","Kiamia@66.com"],["miakia","miakia@77.com"]]
    for row in data:
        sheet.append(row)
    wob.save(filepath)
    print("Done da")
def read():
    rr=load_workbook(filepath)
    sheet=rr.active
    for row in sheet.iter_rows(values_only=True):
        print(row)
def append():
    aa=load_workbook(filepath)
    sheet=aa.active
    data=[["Rey mysterio","Reymysterio+1@g.com"],["Brocklensar","Brock@g.com"]]
    for rows in data:
        sheet.append(rows)
    aa.save(filepath)
write()
append()
read()
