import os
import csv
filepath="C:\\Users\\keerthivasan.a\\python programming\\Book_1.csv"
with open (filepath,"r") as filereader:
    filereaders=csv.DictReader(filereader)
    for files in filereaders:
        print("{},{},{}".format(files["name"],files["username"],files["department"]))
filepath_1="C:\\Users\\keerthivasan.a\\python programming\\Book_19.csv"
os.path.exists(filepath_1)
data = [
    ["Name", " Age", " City"], 
    ["Alice","34", " New York"],
    ["Bob", " 25", " Los Angeles"],
    ["Charlie", " 35", " San Francisco"]]
with open(filepath_1,"a",newline="") as biased:
    bias=csv.writer(biased)
    bias.writerows(data)
filepath_1="C:\\Users\\keerthivasan.a\\python programming\\Book_9.csv"
os.path.exists(filepath_1)
data = [
    ["Name", " Age", " City"], 
    ["Alice","34", " New York"],
    ["Bob", " 25", " Los Angeles"],
    ["Charlie", " 35", " San Francisco"]]
with open(filepath_1,"a+",newline="") as no_header:
    header=csv.writer(no_header)
    header.writerows(data)
    no_header.seek(0)
    print(no_header.read())
filepath_2="C:\\Users\\keerthivasan.a\\python programming\\Book_10.csv"
data=[{"Name":"Aisha Khan", "Department":"Engineering", "Salary":"80000"},{"Name":"Jules Lee", "Department":"Marketing", "Salary":"67000"},
      {"Name":"Queenie Corbit", "Department":"Human Resources", "Salary":"90000"}]
keys=["Name","Department","Salary"]
csv.register_dialect("Dialact",delimiter=",",quotechar="~",escapechar="\\",skipinitialspace=True)
with open(filepath_2,"w+",newline="") as rewrite:
    new_rewrite=csv.DictWriter(rewrite, dialect="Dialact",fieldnames=keys)
    new_rewrite.writeheader()
    new_rewrite.writerows(data)
    rewrite.seek(0)
    print(rewrite.read())


import csv
import os

fileq = "C:\\Users\\keerthivasan.a\\python programming\\append2.csv"

csv.register_dialect("Dialacts", delimiter=",", strict=True, skipinitialspace=True)

datas = [{"Name": "Keerthivasan", "Email": "kkk@gmail.com"}]
keys1 = ["Name", "Email"]

# Open in read+write mode to check if headers exist
file_empty = not os.path.exists(fileq) or os.stat(fileq).st_size == 0

with open(fileq, "a", newline="") as tt:
    ttt = csv.DictWriter(tt, dialect="Dialacts", fieldnames=keys1)
    
    # Write header only if the file is empty
    if file_empty:
        ttt.writeheader()

    # Write data
    ttt.writerows(datas)

print("Data successful")

#xcel
from openpyxl import load_workbook,Workbook
filepath="C:\\Users\\keerthivasan.a\\python programming\\own_Excel1.xlsx"
def writing():
    wob=Workbook()
    sheet=wob.active
    sheet.title="Waste"
    data=[["name","Email"],["Kiamia","Kiamia@66.com"],["miakia","miakia@77.com"]]
    for row in data:
        sheet.append(row)
    wob.save(filepath)
    print("Done")
def read():
    rod=load_workbook(filepath)
    sheets=rod.active
    for row in sheets.iter_rows(values_only=True):
        print(row)
    print("reading done")
def appends():
    loap=load_workbook(filepath)
    shit=loap.active
    data=[["Rey mysterio","Reymysterio+1@g.com"],["Brocklensar","Brock@g.com"]]
    for rwes in data:
        shit.append(rwes)
    loap.save(filepath)
    print("append done")
def delete():
    dee=load_workbook(filepath)
    ow=dee.active
    ow.delete_rows(2)
    dee.save(filepath)
    print("delete")
writing()
read()
appends()
read()
delete()
read()




